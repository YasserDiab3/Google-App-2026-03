"""
Excel to SQLite Migration Engine for HSE System (With Header Deduplication)
"""
import openpyxl
import sqlite3
import os
import json
import datetime
import sys

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\YasserMohamed\Downloads\V.3-HSE Database.xlsx"
DB_PATHS = [
    os.path.abspath(r"backend-sql\data\clinic_hse.db"),
    os.path.abspath(r"Frontend\api\data\clinic_hse.db"),
    os.path.abspath(r"vercel-deploy\api\data\clinic_hse.db")
]

def format_cell_value(val):
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.isoformat()
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False)
    return str(val)

def deduplicate_headers(raw_headers):
    seen = {}
    clean_headers = []
    for idx, h in enumerate(raw_headers):
        val = str(h).strip() if h is not None else ""
        if not val or val.lower() == "none":
            val = f"column_{idx + 1}"
        if val in seen:
            seen[val] += 1
            clean_headers.append(f"{val}_{seen[val]}")
        else:
            seen[val] = 0
            clean_headers.append(val)
    return clean_headers

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at: {EXCEL_PATH}")
        return

    print("====================================================")
    print(f"Starting Migration from: {os.path.basename(EXCEL_PATH)}")
    print("====================================================")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Found {len(sheet_names)} sheets in Excel workbook.\n")

    primary_db_path = DB_PATHS[0]
    os.makedirs(os.path.dirname(primary_db_path), exist_ok=True)
    conn = sqlite3.connect(primary_db_path)
    cursor = conn.cursor()

    total_sheets_imported = 0
    total_records_imported = 0

    for name in sheet_names:
        sheet = wb[name]
        rows_iter = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter, None)
        except StopIteration:
            continue

        if not header_row:
            continue

        headers = deduplicate_headers(header_row)
        if not headers:
            continue

        table_name = name.strip()

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        table_exists = cursor.fetchone()

        if not table_exists:
            col_defs = ", ".join([f'"{h}" TEXT' for h in headers])
            cursor.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({col_defs})')
        else:
            cursor.execute(f'PRAGMA table_info("{table_name}")')
            existing_cols = {col[1] for col in cursor.fetchall()}
            for h in headers:
                if h not in existing_cols:
                    try:
                        cursor.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{h}" TEXT')
                    except Exception:
                        pass

        cursor.execute(f'DELETE FROM "{table_name}"')

        records = []
        for row in rows_iter:
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue
            row_vals = [format_cell_value(row[i]) if i < len(row) else None for i in range(len(headers))]
            records.append(row_vals)

        if records:
            col_names = ", ".join([f'"{h}"' for h in headers])
            placeholders = ", ".join(["?" for _ in headers])
            insert_sql = f'INSERT INTO "{table_name}" ({col_names}) VALUES ({placeholders})'
            cursor.executemany(insert_sql, records)
            conn.commit()

            count = len(records)
            total_sheets_imported += 1
            total_records_imported += count
            print(f"  [OK] [{table_name}]: {count} rows imported")

    conn.close()
    wb.close()

    for other_path in DB_PATHS[1:]:
        os.makedirs(os.path.dirname(other_path), exist_ok=True)
        import shutil
        shutil.copy2(primary_db_path, other_path)
        print(f"  [SYNC] Replicated database to: {os.path.basename(os.path.dirname(os.path.dirname(other_path)))}")

    print("\n====================================================")
    print("MIGRATION COMPLETED SUCCESSFULLY!")
    print(f"   - Sheets Processed: {len(sheet_names)}")
    print(f"   - Sheets with Data: {total_sheets_imported}")
    print(f"   - Total Records:   {total_records_imported}")
    print("====================================================")

if __name__ == "__main__":
    main()
